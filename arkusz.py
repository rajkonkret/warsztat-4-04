import openpyxl

workbook = openpyxl.load_workbook('sales.xlsx')

worksheet = workbook.active

lista = []
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end='\t\t')
        lista.append(col[i].value)
    print('')

print(lista)
print(lista[0:3][0])